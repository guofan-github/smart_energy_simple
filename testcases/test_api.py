# 模块导入
import json
import logging
import openpyxl
import pytest
import requests
from openpyxl.styles import PatternFill

from common.mysql_operate import main
from common.public_api import token_save, send_request, get_data, get_title, get_variable, find_value_by_key, \
    get_variable_dict, login_data_json, get_routerpath, get_variable_dict_by_key, judge_variable, \
    send_request_form_data, send_request_upload
from conf.FC_mqtthub import publish_FC
from conf.M5_mqtthub import publish
from conf.M5_mqtthub_energy import publish_energy
from conf.M5_mqtthub_fault import publish_fault
from conf.M5_mqtthub_warning import publish_warning
from conf.exceldata import Read, backup_excel_file
from conf.公共信息 import path, ip, excel_name

# 配置 logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# 测试类定义
class TestApi:
    data_list = list(get_data())
    json_data = login_data_json()
    result = {}

    # 测试准备
    def setup_class(self):
        global routerpath_dict
        backup_excel_file()
        logging.info('******************************文件已备份*************************************')
        Read().clear_excel_result()
        logging.info('**************************历史测试结果已清除**********************************')
        logging.info('++++++++++++++++++++++++++++++++++测试开始++++++++++++++++++++++++++++++++++')
        token_save()
        routerpath_dict = get_routerpath(self.json_data, self.result)

    # 测试结束
    def teardown_class(self):
        # main()
        # print('****************************已从数据库将新建项目删除**********************************')
        # logging.info('****************************已从数据库将新建项目删除**********************************')
        pass

    # 参数化测试用例
    @pytest.mark.parametrize('data', data_list, ids=get_title())
    def test_api(self, data):
        self.id = int(data[0])
        self.sheet = data[1]
        if data[11] == '否':
            print('跳过该条测试用例')
            self._write_to_excel(self.id, '跳过用例', self.sheet)
            pytest.skip('跳过该条测试用例')
        self._prepare_test_data(data)
        res = self.send_request_based_on_type()
        self._assert_response(res, data)
        self._mqtt_publish()

    def send_request_based_on_type(self):
        if self.json2 is not None and 'form-data' in self.json2:
            print('发送form_data请求')
            res = send_request_form_data(self.url, self.json1, self.headers)
        elif self.json2 is not None and '[' in self.json2 and '{' not in self.json2 and '&' not in self.json2:
            # 针对传参只有一个列表的post请求
            print('发送json只有一个列表的post请求')
            res2 = requests.post(url=self.url, json=self.json1, headers=self.headers)
            res = json.loads(res2.text)
        elif self.json2 == "test_uploadfile.xlsx":
            print('文件上传接口')
            res = send_request_upload(self.url, self.headers, self.json2)
        elif self.json2 == 'test_picture.jpg':
            print('文件上传接口')
            res = send_request_upload(self.url, self.headers, self.json2)
        else:
            print('正常请求')
            res = send_request(self.method, self.url, self.json1, self.headers)

        logging.info(f'发送请求中的方法为：{self.method},url为：{self.url},参数为：{self.json1},请求头为：{self.headers}')
        logging.info(f'接口返回结果为：{res}')
        return res

    # 准备测试数据
    def _prepare_test_data(self, data):
        self.id = int(data[0])
        logging.info(f'接口传入的id为：{self.id}')
        self.sheet = data[1]
        logging.info(f'当前sheet为：{self.sheet}')
        self.module = data[2]
        logging.info(f'接口传入的模块为：{self.module}')
        self.name = data[3]
        logging.info(f'接口传入的接口名称为：{self.name}')
        self.desc = data[4]
        logging.info(f'接口传入的描述为：{self.desc}')
        self.method = data[5].lower()
        logging.info(f'接口传入的请求方式为：{self.method}')
        self.url = f'{ip}{data[6]}'
        logging.info(f'接口传入的url为：{self.url}')
        self.json2 = data[7]
        print(f'接口传入的id为：{self.id},当前sheet为：{self.sheet},模块为：{self.module},接口名称为：{self.name},'
              f'描述为：{self.desc},请求方式为：{self.method},url为：{self.url},json为：{self.json2}')
        if data[7]:
            print('接口传入的json不为空')
            if 'form-data' in data[7]:
                print('接口传入的json中包含form-data')
                self.json1 = data[7].split('=')[1]
                logging.info(f'接口传入的form-data为：{self.json1},type为{type(self.json1)}')
            elif '[' in data[7] and '{' not in data[7] and '&' not in data[7]:
                print('接口传入的json中只包含列表')
                self.json1 = json.loads(data[7])
                logging.info(f'接口传入的单独列表为：{self.json1}')
            elif "test_uploadfile.xlsx" in data[7]:
                print('文件上传接口传参')
                self.json1 = data[7]
                logging.info(f'接口传入的文件上传参数为：{self.json1}')
            else:
                print('接口正常传入json')
                self.json1 = data[7]
                logging.info(f'接口传入的json为：{self.json1}')
        else:
            print('接口传入的json为空')
            self.json1 = None
            logging.info(f'接口传入的json为：{self.json1}')
        self.headers = {'Authorization': f'{token_save()}'}
        module = self.module
        if module in routerpath_dict:
            print(f'模块routerpath字典：{routerpath_dict}')
            route_url = routerpath_dict[module]
            logging.info(f'找到模块对应的routeUrl：{route_url}')
            self.headers['Router-Path'] = route_url
            logging.info(f'接口传入的请求头为：{self.headers}')
        else:
            logging.warning(f'未找到模块{module}对应的routeUrl')

        if data[9]:
            print('接口传入的使用变量不为空')
            self.json1 = judge_variable(self.method, self.json1, data[9])
            logging.info(f'替换后的传参为:{self.json1}')

        if data[8]:
            print('接口传入的获取变量不为空')
            if data[7] == 'test_picture.jpg' or data[7] == 'test_uploadfile.xlsx':
                dict_res = send_request_upload(self.url, self.headers, self.json2)
            else:
                dict_res = get_variable(self.method, self.url, self.json1, self.headers)
            logging.info(f'获取到做变量的值为：{data[8]}')
            logging.info(f'接口返回为：{dict_res}')
            value = find_value_by_key(dict_res, data[8])
            logging.info(f'获取到的值为：{value}')
            get_variable_dict(data[8], value)

        logging.info('测试数据准备完毕')

    # 断言响应结果
    def _assert_response(self, res, data):
        res_ = json.dumps(res, ensure_ascii=False)
        print(f'接口返回结果为：{res_}')
        expected = str(data[10])
        if get_variable_dict_by_key(expected):
            print('期望结果为变量')
            expected = get_variable_dict_by_key(expected)
        logging.info(f'期望结果为：{expected}')

        try:
            assert expected in res_
            self.results = '测试通过'
            logging.info('断言成功：测试通过')
        except Exception as e:
            logging.error(f'断言失败，失败原因为：{e}')
            self.results = f'测试失败:{str(e)}'
            raise Exception('断言失败')
        finally:
            logging.info(f'写回测试结果到excel：{self.results}')
            self._write_to_excel(self.id, self.results, self.sheet)

    # 将测试结果写入excel
    def _write_to_excel(self, id, results, sheetname):
        print(f'写回测试结果到excel：{results}')
        file = openpyxl.load_workbook(f'{path}{excel_name}.xlsx')
        sheet = file[sheetname]
        cell = sheet.cell(id + 1, 12)
        cell.value = results
        if results == '测试通过':
            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        if '测试失败' in results:
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        if results == '跳过用例':
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

        file.save(f'{path}{excel_name}.xlsx')

    # 测试数据上报
    def _mqtt_publish(self):
        if self.id == 2 and self.module == '电力监测':
            publish()
            print('上报第一包能耗报文')
            logging.info('*************************mqtt测试数据已发送*******************************')
        if self.id == 1 and self.module == '故障告警':
            publish_fault()
            print('上报故障报文')
            logging.info('*************************故障测试数据已发送*******************************')
        if self.id == 3 and self.module == '电力监测':
            publish_energy()
            print('上报第二包能耗报文')
            logging.info('*************************能耗测试数据已发送*******************************')
        if self.id == 2 and self.module == '故障告警':
            publish_warning()
            print('上报告警报文')
            logging.info('*************************告警测试数据已发送*******************************')
        if self.id % 25 == 0:
            publish_FC()
            print('上报粉尘传感器数据')
            logging.info('*************************粉尘传感器数据已发送******************************')


# 执行测试
if __name__ == '__main__':
    pytest.main(['-v', 'test_api.py'])
