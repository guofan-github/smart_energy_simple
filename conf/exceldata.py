import os
import shutil

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

from conf.公共信息 import path, excel_name, ui_data_path


def backup_excel_file():
    file_path = path+excel_name+'.xlsx'
    # 如果文件不存在，则无需备份
    if not os.path.exists(file_path):
        print("文件不存在，无法备份。")
        return

    # 获取文件名和扩展名
    file_name, file_extension = os.path.splitext(file_path)

    # 构造备份文件名，例如：original_file.xlsx -> original_file_backup.xlsx
    backup_file_path = f"{file_name}_backup{file_extension}"

    try:
        # 备份文件
        shutil.copyfile(file_path, backup_file_path)
        print(f"成功备份文件：{file_path} -> {backup_file_path}")
    except Exception as e:
        print(f"备份文件失败：{e}")


class Read:
    def __init__(self):
        # 加载 Excel 文件
        self.wb = load_workbook(f'{path}{excel_name}.xlsx')
        self.sheet = None

    def get_data(self):

        data_list = []
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            rows = ws.max_row
            cols = ws.max_column
            for r in range(2, rows + 1):
                lines = []
                for c in range(1, cols + 1):
                    lines.append(ws.cell(r, c).value)
                data_list.append(lines)

        self.wb.close()
        return data_list

    def clear_excel_result(self):

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            # 指定要操作的列和范围
            column_index = 12  # 例如第一列
            start_row = 2  # 从第二行开始

            # border = Border(
            #     left=Side(border_style='thin', color='000000'),
            #     right=Side(border_style='thin', color='000000'),
            #     top=Side(border_style='thin', color='000000'),
            #     bottom=Side(border_style='thin', color='000000')
            # )

            # 设置填充颜色
            fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白色填充

            # 遍历指定列的单元格
            for row in ws.iter_rows(min_row=start_row, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.value = None
                    cell.fill = fill
                    # cell.border = border

            # 保存文件
            self.wb.save(f'{path}{excel_name}.xlsx')


def backup_excel_file_ui():
    file_path = ui_data_path
    # 如果文件不存在，则无需备份
    if not os.path.exists(file_path):
        print("文件不存在，无法备份。")
        return

    # 获取文件名和扩展名
    file_name, file_extension = os.path.splitext(file_path)

    # 构造备份文件名，例如：original_file.xlsx -> original_file_backup.xlsx
    backup_file_path = f"{file_name}_backup{file_extension}"

    try:
        # 备份文件
        shutil.copyfile(file_path, backup_file_path)
        print(f"成功备份文件：{file_path} -> {backup_file_path}")
    except Exception as e:
        print(f"备份文件失败：{e}")


class ReadUi:
    def __init__(self):
        # 加载 Excel 文件
        self.wb = load_workbook(ui_data_path)
        self.sheet = None

    def get_data_ui(self):

        # 初始化用例列表
        test_cases = []

        # 遍历每个sheet读取数据
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]

            # 标题
            data = sheet.cell(row=1, column=1).value

            # 提取数据并组织成用例
            current_case = None
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if row[1]:  # 如果用例标题不为空，表示一个新的用例开始
                    if current_case:  # 如果当前用例不为空，将其加入到用例列表中
                        test_cases.append(current_case)
                    current_case = {
                        'sheet标题': data,
                        '用例标题': row[1],
                        '用例步骤': [],
                        '是否执行': row[8]  # 初始化用例步骤列表
                    }
                # 将当前行作为用例的步骤添加到用例步骤列表中
                current_case['用例步骤'].append({
                    '序号': row[0],
                    '步骤描述': row[2],
                    '关键字行为': row[3],
                    '定位方法': row[4],
                    '元素位置': row[5],
                    '输入内容': row[6],
                    '断言值': row[7]
                })

            # 将最后一个用例加入到用例列表中
            if current_case:
                test_cases.append(current_case)

        return test_cases

    def clear_excel_result_ui(self):
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            # 指定要操作的列和范围
            column_index = 10  # 例如第一列
            start_row = 3  # 从第二行开始

            # 设置填充颜色
            fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白色填充
            # 遍历指定列的单元格
            for row in ws.iter_rows(min_row=start_row, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.value = None
                    cell.fill = fill
                    # cell.border = border

            # 保存文件
            self.wb.save(ui_data_path)

    # 将测试结果写入excel
    def _write_to_excel(self, id, results, sheetname):
        sheet = self.wb[sheetname]
        cell = sheet.cell(id + 2, 10)
        cell.value = results
        if results == '测试通过':
            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        if '测试失败' in results:
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        if results == '跳过用例':
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

        self.wb.save(ui_data_path)


if __name__ == '__main__':
    backup_excel_file_ui()


